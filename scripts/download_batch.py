import os
import requests
import certifi
import urllib3
from datetime import datetime
import openpyxl

# Disable SSL warnings when verify=False is used
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

RAW_DIR = r"C:\Users\vaish\OneDrive\Desktop\sahayak\data\raw"
SOURCE_LIST = r"C:\Users\vaish\OneDrive\Desktop\sahayak\data\metadata\sources_to_download.xlsx"
META_FILE = r"C:\Users\vaish\OneDrive\Desktop\sahayak\data\metadata\source_registry.xlsx"

os.makedirs(RAW_DIR, exist_ok=True)

def download_and_log(source_name, url, filename, ws):
    filepath = os.path.join(RAW_DIR, filename)

    # Skip if file already exists
    if os.path.exists(filepath):
        print(f"⏩ Skipping {filename}, already exists in raw folder")
        return

    print(f"Downloading {filename} from {url}...")
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        }

        # Conditional SSL bypass for Gazette
        if "egazette.gov.in" in url:
            response = requests.get(url, headers=headers, verify=False)
        else:
            response = requests.get(url, headers=headers, verify=certifi.where())

        response.raise_for_status()

        # Detect file type from Content-Type header
        content_type = response.headers.get("Content-Type", "").lower()
        if "pdf" in content_type and not filename.endswith(".pdf"):
            filename += ".pdf"
            filepath = os.path.join(RAW_DIR, filename)
        elif "html" in content_type and not filename.endswith(".html"):
            filename += ".html"
            filepath = os.path.join(RAW_DIR, filename)

        with open(filepath, "wb") as f:
            f.write(response.content)

        date_downloaded = datetime.today().strftime("%Y-%m-%d")
        ws.append([source_name, filename, url, date_downloaded])
        print(f"✅ Downloaded {filename} and logged in source_registry.xlsx")

    except Exception as e:
        print(f"⚠️ Failed to download {filename} from {url}: {e}")

# Load source list
print("Reading sources_to_download.xlsx...")
wb_sources = openpyxl.load_workbook(SOURCE_LIST)
sheet = wb_sources.active

# Prepare registry workbook
if os.path.exists(META_FILE):
    wb_registry = openpyxl.load_workbook(META_FILE)
    ws_registry = wb_registry.active
else:
    wb_registry = openpyxl.Workbook()
    ws_registry = wb_registry.active
    ws_registry.append(["source_name", "file_name", "url", "date_downloaded"])

# Process rows
for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
    source_name, url, filename = row
    if source_name and url and filename:
        download_and_log(source_name, url, filename, ws_registry)

# Save registry
wb_registry.save(META_FILE)
print("🎯 All downloads complete.")
